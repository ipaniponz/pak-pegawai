"""Migrasi IDENTITAS pegawai dari spreadsheet "Monitoring AK" (format lama) ke
skema baru (pegawai, riwayat_jenjang). TIDAK mengimpor angka kredit historis --
sesuai keputusan: AK awal tiap pegawai diisi MANUAL oleh admin lewat form
"Catat Naik Jenjang" di aplikasi setelah migrasi (ak_awal_jenjang selalu 0 di
sini), bukan ditebak dari kolom "Nilai PAK 2024"/"Nilai 2025"/"Nilai 2026 (Jan-Mar)"
yang formatnya tidak konsisten (lihat kasus "Belum Dihitung").

Cara pakai:
    python scripts/migrate_spreadsheet.py "1. Angka Kredit JF Biro Hukum dan Estimasi Kenaikan (1).xlsx"

Output:
    seed/seed_data.sql  -> INSERT pegawai + riwayat_jenjang (AK awal = 0, sk_referensi
                           menjelaskan ini hasil migrasi identitas, bukan AK final)
    flagged_rows.csv     -> baris yang jabatannya tidak bisa dipetakan ke jenjang manapun,
                           perlu direview manual oleh admin

Prinsip migrasi (PENTING, jangan diubah tanpa sadar konsekuensinya):
  - HANYA identitas (nama, jabatan_fungsional, substansi, kategori_jf, jenjang
    sekarang, status_kepegawaian) yang diimpor otomatis.
  - Kolom "Nilai PAK 2024", "Nilai 2025", "Nilai 2026 (Jan-Mar)" SENGAJA TIDAK
    diimpor -- admin akan input AK awal tiap pegawai secara manual lewat
    aplikasi (form "Catat Naik Jenjang"), karena nilai-nilai itu campuran
    format lama yang sebagian tidak valid (lihat kasus "Belum Dihitung").
  - jabatan_fungsional dipetakan ke nama dasar resmi (tanpa jenjang) sesuai
    PETA_JABATAN di bawah -- JANGAN simpan teks mentah dari sumber (lihat 1.4m
    AGENT_BRIEF: kalau dibiarkan gabung dgn jenjang, lookup tembusan otomatis
    akan selalu kosong).
  - Kolom "Jenjang Target" bernilai 'CPNS' bukan nama jenjang -- itu sinyal
    status_kepegawaian='CPNS', BUKAN jenjang yang harus diparse.
"""
import sys
import csv
from openpyxl import load_workbook

JENJANG_KEAHLIAN = ["Ahli Pertama", "Ahli Muda", "Ahli Madya", "Ahli Utama"]
JENJANG_KETERAMPILAN = ["Pemula", "Terampil", "Mahir", "Penyelia"]

# Nama dasar resmi jabatan fungsional (tanpa jenjang) -- lihat 1.4m AGENT_BRIEF.
# HARUS sama persis dengan jabatan_fungsional yang dipakai di seed/tembusan_referensi.sql
# supaya lookup tembusan otomatis tidak kosong.
PETA_JABATAN = {
    "Perancang PUU": "Perancang Peraturan Perundang-undangan",
    "Analis Hukum": "Analis Hukum",
    "Perencana": "Perencana",
    "Analis SDMA": "Analis SDM Aparatur",
    "Arsiparis": "Arsiparis",
}


def parse_jabatan(jabatan: str):
    """Pecah 'Perancang PUU Ahli Madya' -> ('Perancang Peraturan Perundang-undangan', 'Ahli Madya', 'Keahlian')."""
    for j in JENJANG_KEAHLIAN:
        if jabatan.endswith(j):
            prefix = jabatan[: -len(j)].strip()
            return PETA_JABATAN.get(prefix, prefix), j, "Keahlian"
    for j in JENJANG_KETERAMPILAN:
        if jabatan.endswith(j):
            prefix = jabatan[: -len(j)].strip()
            return PETA_JABATAN.get(prefix, prefix), j, "Keterampilan"
    return jabatan, None, None


def sql_str(val):
    if val is None:
        return "NULL"
    return "'" + str(val).replace("'", "''") + "'"


def main(path):
    wb = load_workbook(path, data_only=True)
    ws = wb["Monitoring AK"]

    seed_lines = []
    flagged = []
    pegawai_id = 0

    for row in range(5, ws.max_row + 1):
        nama = ws.cell(row=row, column=3).value
        jabatan = ws.cell(row=row, column=4).value
        jenjang_target = ws.cell(row=row, column=5).value

        if nama is None or jabatan is None:
            continue  # baris kosong / header section

        pegawai_id += 1
        nama = str(nama).strip()

        if jenjang_target == "Struktural":
            seed_lines.append(
                f"INSERT INTO pegawai (id, nip, nama, kategori_jf, jabatan_fungsional, "
                f"substansi, status, data_lengkap, status_kepegawaian) VALUES "
                f"({pegawai_id}, NULL, {sql_str(nama)}, 'Struktural', {sql_str(jabatan)}, "
                f"NULL, 'aktif', 0, 'PNS');"
            )
            continue

        nama_dasar, jenjang_sekarang, kategori_jf = parse_jabatan(str(jabatan))
        if jenjang_sekarang is None:
            flagged.append({
                "row": row,
                "nama": nama,
                "issue": f"Tidak bisa parse jenjang dari Jabatan='{jabatan}' -- review manual.",
            })
            continue

        status_kepegawaian = "CPNS" if jenjang_target == "CPNS" else "PNS"
        substansi = str(jabatan)[: -len(jenjang_sekarang)].strip()

        seed_lines.append(
            f"INSERT INTO pegawai (id, nip, nama, kategori_jf, jabatan_fungsional, "
            f"substansi, status, data_lengkap, status_kepegawaian) VALUES "
            f"({pegawai_id}, NULL, {sql_str(nama)}, {sql_str(kategori_jf)}, "
            f"{sql_str(nama_dasar)}, {sql_str(substansi)}, 'aktif', 0, {sql_str(status_kepegawaian)});"
        )
        seed_lines.append(
            f"INSERT INTO riwayat_jenjang (pegawai_id, jenjang_referensi_id, "
            f"tanggal_mulai, tanggal_selesai, ak_awal_jenjang, sk_referensi) VALUES "
            f"({pegawai_id}, (SELECT id FROM jenjang_referensi WHERE kategori="
            f"{sql_str(kategori_jf)} AND nama_jenjang={sql_str(jenjang_sekarang)}), "
            f"'2024-01-01', NULL, 0, "
            f"'Migrasi identitas dari spreadsheet lama -- AK awal BELUM diisi, "
            f"admin wajib input manual setelah migrasi');"
        )

    with open("seed/seed_data.sql", "w", encoding="utf-8") as f:
        f.write("-- Auto-generated oleh scripts/migrate_spreadsheet.py. Review flagged_rows.csv dulu.\n")
        f.write("-- AK awal SENGAJA 0 untuk semua baris -- admin wajib isi manual lewat aplikasi.\n")
        f.write("\n".join(seed_lines) + "\n")

    with open("flagged_rows.csv", "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["row", "nama", "issue"])
        writer.writeheader()
        writer.writerows(flagged)

    print(f"Selesai. {pegawai_id} pegawai diproses, {len(seed_lines)} statement INSERT, "
          f"{len(flagged)} baris di-flag untuk review manual.")
    print("-> seed/seed_data.sql")
    print("-> flagged_rows.csv")
    print("Ingat: AK awal tiap pegawai (ak_awal_jenjang) masih 0 -- isi manual lewat "
          "form 'Catat Naik Jenjang' di halaman detail pegawai setelah data ini diimpor.")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python scripts/migrate_spreadsheet.py <path_to_xlsx>")
        sys.exit(1)
    main(sys.argv[1])
